#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
# -----------------------------------------------------------------------------
# Purpose:    transfer xml string to Excel model
# Author:      lianjiaxin
# Created:     2017-09-25
# Copyright:   (c) Hikvision.com 2016
# -----------------------------------------------------------------------------
'''
import sys
import xlsxwriter
from openpyxl import load_workbook
from configFile import *
reload(sys)
sys.setdefaultencoding('utf-8')

class readexcel():
    """将XML格式的分辨率数据转换为Excel数据"""
    def __init__(self, file_path=None):
        self.qiandaQuestionList = []
        self.fengxianQuestionList = []
        self.lundaQuestionList = []
        self.srcFile = file_path
        print "cfgpath:",self.srcFile
        self.configFile = configFile()

    def readExcelQiangda(self):
        self.qiandaQuestionList = []
        sheetName = '抢答题'
        try:
            #源表格:加载源表格，默认可读写
            srcWb = load_workbook(self.srcFile)
            #源表格:根据sheet页名称获得sheet
            srcWs = srcWb[sheetName]
        except Exception as e:
            print ("exception happen:",e)
        else:
            
            #遍历i= 1,2，行索引
            for i in range(2,srcWs.max_row+1):
                #遍历j= 1，列索引
                #源表格：获取源表格单元格A1，A2的内容，读取行列索引从1开始
                cfgIdx = srcWs.cell(row = i,column=1)
                answer=srcWs.cell(row = i,column=2)
                questype = srcWs.cell(row = i,column=3)
                content = srcWs.cell(row = i,column=4)
                if questype.value== '问答题':
                    quesstr = content.value
                else:
                    dA = srcWs.cell(row = i,column=5)
                    dB = srcWs.cell(row = i,column=6)
                    dC = srcWs.cell(row = i,column=7)
                    dD = srcWs.cell(row = i,column=8)
                    #print dA.value,dB.value,dC.value,dD.value
                    quesstr = "<br>A."+str(dA.value)+"<br>B."+str(dB.value)+"<br>C."+str(dC.value)+"<br>D."+str(dD.value)
                    quesstr = content.value+quesstr
                dict={'type':questype.value,'content':quesstr,'answer':answer.value,'cfgIdx':str(cfgIdx.value),'sheet':sheetName}
                self.qiandaQuestionList.append(dict)
            print "qiangda:",len(self.qiandaQuestionList)
            cfgIdxList = self.configFile.readCfg(sheetName)
            for i in cfgIdxList:
                print 'len',len(self.qiandaQuestionList)
                for j in range(0,len(self.qiandaQuestionList)):
                    #print j,self.qiandaQuestionList[j]['cfgIdx']
                    if i == self.qiandaQuestionList[j]['cfgIdx']:
                        #print 'a',i,self.qiandaQuestionList[j]['cfgIdx']
                        del self.qiandaQuestionList[j]
                        break
            print "qiangda:",len(self.qiandaQuestionList)


    def readExcelFengxian(self):
        self.fengxianQuestionList = []
        sheetName = '风险题'
        try:
            #源表格:加载源表格，默认可读写
            srcWb = load_workbook(self.srcFile)
            #源表格:根据sheet页名称获得sheet
            srcWs = srcWb[sheetName]
        except Exception as e:
            print ("exception happen:",e)
        else:
            
            #遍历i= 1,2，行索引
            for i in range(2,srcWs.max_row+1):
                #遍历j= 1，列索引
                #源表格：获取源表格单元格A1，A2的内容，读取行列索引从1开始
                cfgIdx = srcWs.cell(row = i,column=1)
                hardLevel=srcWs.cell(row = i,column=2)
                quesSub= srcWs.cell(row = i,column=3)
                answer = srcWs.cell(row = i,column=4)
                questype= srcWs.cell(row = i,column=5)
                content = srcWs.cell(row = i,column=6)
                if questype.value== '问答题':
                    quesstr = content.value
                else:
                    dA = srcWs.cell(row = i,column=7)
                    dB = srcWs.cell(row = i,column=8)
                    dC = srcWs.cell(row = i,column=9)
                    dD = srcWs.cell(row = i,column=10)
                    #print dA.value,dB.value,dC.value,dD.value
                    quesstr = "<br>A."+str(dA.value)+"<br>B."+str(dB.value)+"<br>C."+str(dC.value)+"<br>D."+str(dD.value)
                    quesstr = content.value+quesstr
                dict={'type':questype.value,'content':quesstr,'answer':answer.value,'level':str(hardLevel.value),'subject':quesSub.value,'cfgIdx':str(cfgIdx.value),'sheet':sheetName}
                self.fengxianQuestionList.append(dict)
            print "fengxian:",len(self.fengxianQuestionList)
            cfgIdxList = self.configFile.readCfg(sheetName)
            for i in cfgIdxList:
                for j in range(0,len(self.fengxianQuestionList)):
                    if i == self.fengxianQuestionList[j]['cfgIdx']:
                        del self.fengxianQuestionList[j]
                        break
            print "fengxian:",len(self.fengxianQuestionList)

    def readExcellunda(self):
        self.lundaQuestionList = []
        sheetName = '轮答题'
        try:
            #源表格:加载源表格，默认可读写
            srcWb = load_workbook(self.srcFile)
            #源表格:根据sheet页名称获得sheet
            srcWs = srcWb[sheetName]
        except Exception as e:
            print ("exception happen:",e)
        else:
            
            #遍历i= 1,2，行索引
            for i in range(2,srcWs.max_row+1):
                #遍历j= 1，列索引
                #源表格：获取源表格单元格A1，A2的内容，读取行列索引从1开始
                cfgIdx = srcWs.cell(row = i,column=1)
                idx=srcWs.cell(row = i,column=2)
                answer = srcWs.cell(row = i,column=3)
                questype= srcWs.cell(row = i,column=4)
                content = srcWs.cell(row = i,column=5)
                if questype.value== '问答题':
                    quesstr = content.value
                else:
                    dA = srcWs.cell(row = i,column=6)
                    dB = srcWs.cell(row = i,column=7)
                    dC = srcWs.cell(row = i,column=8)
                    dD = srcWs.cell(row = i,column=9)
                    #print dA.value,dB.value,dC.value,dD.value
                    quesstr = "<br>A."+str(dA.value)+"<br>B."+str(dB.value)+"<br>C."+str(dC.value)+"<br>D."+str(dD.value)
                    quesstr = content.value+quesstr
                dict={'type':questype.value,'content':quesstr,'answer':answer.value,'idx':idx.value,'cfgIdx':str(cfgIdx.value),'sheet':sheetName}
                self.lundaQuestionList.append(dict)
            print "lunda:",len(self.lundaQuestionList)
            cfgIdxList = self.configFile.readCfg(sheetName)
            for i in cfgIdxList:
                for j in range(0,len(self.lundaQuestionList)):
                    if i == self.lundaQuestionList[j]['cfgIdx']:
                        del self.lundaQuestionList[j]
                        break
            print "lunda:",len(self.lundaQuestionList)

if __name__ == "__main__":
    test = readexcel()
    test.readExcelQiangda()
    test.readExcelFengxian()
    test.readExcellunda()
